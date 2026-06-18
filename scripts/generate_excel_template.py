"""
產生 Power Sequence Node 匯入用 Excel 範本（v5 長表 + VBA 同步）。
執行: python scripts/generate_excel_template.py
輸出: templates/powerseq_nodes_template.xlsm

工作表：
  Instructions — how to fill the workbook
  Config       — project defaults
  Nodes        — main sheet: all nodes (row order = system order)
  Output Conditions — one row = one AND group
  Input Conditions  — WaveDrom Input Hi/Lo (Custom wave or Signal cond.)
  Lists        — dropdown sources (hidden)
"""
from __future__ import annotations

import os
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    print("請先安裝: pip install openpyxl")
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_SRC = os.path.join(ROOT, "src")
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)
from excel_import import build_signal_dropdown_entries
from excel_template_layout import apply_nodes_sheet_header_rows, node_sheet_headers
OUT_PATH = os.path.join(ROOT, "templates", "powerseq_nodes_template.xlsm")
TMP_XLSX = os.path.join(ROOT, "templates", "_powerseq_nodes_template_build.xlsx")

SHEET_INSTRUCTIONS = "Instructions"
SHEET_NODES = "Nodes"
SHEET_COND = "Output Conditions"
SHEET_INPUT_COND = "Input Conditions"
SHEET_LISTS = "Lists"
SHEET_LEGEND = "Field Reference"
DATA_START_ROW = 4

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
INPUT_FILL = PatternFill("solid", fgColor="F3E8FF")
OUTPUT_FILL = PatternFill("solid", fgColor="DBEAFE")
HI_FILL = PatternFill("solid", fgColor="D1FAE5")
LO_FILL = PatternFill("solid", fgColor="FEE2E2")
FORCE_FILL = PatternFill("solid", fgColor="FEF3C7")
SECTION_FONT = Font(bold=True, size=11)

SIGNAL_PRESETS = ("High", "Low")
OUTPUT_USE_SUFFIXES = ("|Hi Cond", "|Lo Cond", "|Force Cond")

COND_SIGNAL_INIT_COLS = 8
COND_SIGNAL_MAX_COLS = 49
COND_META_COLS = 4

INPUT_COND_SIGNAL_INIT_COLS = 8
INPUT_COND_SIGNAL_MAX_COLS = 49
INPUT_META_COLS = 6  # input_name, cond_side, mode, wave, operation, group_inv

INPUT_MODES = ("Low (0)", "High (1)", "Custom wave", "Signal cond.")

CONFIG_ROWS = [
    ("key", "Label", "Description", "Example"),
    ("module_name", "module_name", "Module name (optional)", "PWRSEQ_TOP"),
    ("pulses", "pulses", "Pulse list, comma-separated (dropdown source)", "Pulse_1us,Pulse_2ms"),
    ("default_cycle_hi", "default_cycle_hi", "When cycle_hi blank (Output CYCLE_HI / Input DEB CYCLE_HI)", "8"),
    ("default_cycle_lo", "default_cycle_lo", "When cycle_lo blank (Output CYCLE_LO / Input DEB CYCLE_LO)", "4"),
    ("default_cycle_force", "default_cycle_force", "When Output cycle_force blank", "2"),
    ("default_init", "default_init", "When init blank (Output INIT / Input DEB INIT)", "0"),
    ("default_force_val", "default_force_val", "When force_val blank", "0"),
    ("default_pulse", "default_pulse", "When pulse field blank", "Pulse_1us"),
    ("wavedrom_steps", "wavedrom_steps", "WaveDrom simulation steps", "50"),
    ("wavedrom_hscale", "wavedrom_hscale", "WaveDrom hscale (pixels per step)", "1"),
]

# name, type, cycle_hi, cycle_lo, cycle_force, init, force_val, pulse_hi, pulse_lo, pulse_timing
NODE_ROWS = [
    ("EKEY", "Input", 2, 2, None, 0, None, "Pulse_1us", "Pulse_1us", "Pulse_1us", INPUT_FILL),
    ("PRIM_VR_EN", "Input", 0, 0, None, 0, None, "Pulse_1us", "Pulse_1us", "Pulse_1us", INPUT_FILL),
    (
        "PCH_P0V85A_EN", "Output", 5, 4, 2, 0, 0,
        "Pulse_2ms", "Pulse_1us", "Pulse_1us", OUTPUT_FILL,
    ),
]

# (output_name, cond_type, operation, group_inv, *signals)
COND_ROWS = [
    ("PCH_P0V85A_EN", "Hi", "AND", "N", "EKEY", "PRIM_VR_EN", "!RSMRST_N"),
    ("", "Lo", "AND", "N", "Low"),
    ("", "Force", "AND", "N"),
]

# (input_name, side, mode, wave, operation, group_inv, *signals)
INPUT_COND_ROWS = [
    ("EKEY", "Hi", "Custom wave", "01", "AND", "N"),
    ("", "Lo", "Low (0)", None, "AND", "N"),
    ("PRIM_VR_EN", "Hi", "Signal cond.", None, "AND", "N", "EKEY"),
    ("", "Lo", "Low (0)", None, "AND", "N"),
]

INSTRUCTIONS = [
    "Power Sequence Generator — Excel Template",
    "",
    "Workflow",
    "  1. Config: edit pulses (comma-separated) — Sync updates pulse dropdowns on Nodes",
    "  2. Nodes: one row per node (top-to-bottom = system order)",
    "  3. Click Sync Conditions (top-right on Nodes sheet)",
    "  4. Output Conditions: each Output gets 1 Hi + 1 Lo + 1 Force row (add rows for OR)",
    "  5. Input Conditions: each Input gets 1 Hi + 1 Lo row (WaveDrom only, not Verilog)",
    "",
    "Nodes (10 columns)",
    "  name / type",
    "  cycle_hi, cycle_lo: Output → CYCLE; Input → DEB CYCLE (any ≠ 0 enables debounce)",
    "  cycle_force: Output only",
    "  init: Output → INIT; Input → DEB INIT",
    "  force_val: FORCE output 0/1",
    "  pulse_hi, pulse_lo",
    "  pulse_timing: Output → Pulse_Force; Input → Pulse_Deb",
    "",
    "Output Conditions (long table)",
    "  output_name | Cond Type (Hi/Lo/Force) | Operation (AND/OR/XOR) | Group Inv (Y/N) | Signal…",
    "  Multiple signals on one row = group Operation; multiple rows same output + type = OR",
    "  Prefix ! on signal = invert; blank Group Inv = N (0)",
    "  Output may use NAME|Hi Cond / |Lo Cond / |Force Cond (GUI use dropdown)",
    "  NAME only = Node (self)",
    "",
    "Input Conditions (WaveDrom only)",
    "  input_name | Side (Hi/Lo) | Mode | Wave | Operation (AND/OR/XOR) | Group Inv | Signal…",
    "  Mode: Low (0) / High (1) / Custom wave / Signal cond.",
    "  Custom wave: fill Wave with 0/1/. or 0{29}1 (e.g. high at step 30)",
    "  Signal cond.: same syntax as Output (Operation per row, OR across rows); leave Wave empty",
    "  Hi default: Signal cond.; Lo default: Low (0)",
    "  Config wavedrom_steps / wavedrom_hscale used when exporting WaveDrom",
    "",
    "Macros & button",
    "  Enable Content when opening the workbook",
    "  Nodes sheet: Sync Conditions — refreshes Output/Input conditions and Lists",
    "  Or Alt+F8 → RequestSyncFromNodes",
    "",
    "Import",
    "  Open the workbook in PowerSeq Generator (Open → .xlsx / .xlsm)",
]


def _node_headers():
    return node_sheet_headers()


def _cond_headers():
    cols = [
        ("output_name", "output_name", "Same as Output on Nodes; first row of block"),
        ("cond_type", "Cond Type (Hi/Lo/Force)", "Hi / Lo / Force (dropdown)"),
        ("operation", "Operation (AND/OR/XOR)", "AND / OR / XOR within this group"),
        ("group_inv", "Group Inv (Y/N)", "Invert whole group; blank = N (0)"),
    ]
    for i in range(1, COND_SIGNAL_INIT_COLS + 1):
        label = "Signal" if i == 1 else ""
        hint = (
            "Signal dropdown; ! = invert; Output may use NAME|Hi/Lo/Force Cond; insert cols to extend"
            if i == 1
            else ""
        )
        cols.append((f"signal{i}", label, hint))
    return cols


def _cond_total_cols() -> int:
    return COND_META_COLS + COND_SIGNAL_INIT_COLS


def _cond_signal_end_col() -> int:
    return COND_META_COLS + COND_SIGNAL_MAX_COLS


def _input_cond_headers():
    cols = [
        ("input_name", "input_name", "Same as Input on Nodes; first row of block"),
        ("cond_side", "Side (Hi/Lo)", "Hi / Lo (dropdown)"),
        ("mode", "Mode", "Low / High / Custom wave / Signal cond. (dropdown)"),
        ("wave", "Wave", "Custom wave: 0/1/. or 0{n}1"),
        ("operation", "Operation (AND/OR/XOR)", "AND / OR / XOR within this group"),
        ("group_inv", "Group Inv (Y/N)", "Signal cond. group invert; blank = N"),
    ]
    for i in range(1, INPUT_COND_SIGNAL_INIT_COLS + 1):
        label = "Signal" if i == 1 else ""
        hint = (
            "Signal cond.: dropdown; ! = invert; NAME|Hi/Lo Cond allowed"
            if i == 1
            else ""
        )
        cols.append((f"signal{i}", label, hint))
    return cols


def _input_cond_total_cols() -> int:
    return INPUT_META_COLS + INPUT_COND_SIGNAL_INIT_COLS


def _input_cond_signal_end_col() -> int:
    return INPUT_META_COLS + INPUT_COND_SIGNAL_MAX_COLS


def _input_side_fill(side: str | None) -> PatternFill | None:
    if not side:
        return INPUT_FILL
    key = str(side).strip().lower()
    if key == "hi":
        return HI_FILL
    if key == "lo":
        return LO_FILL
    return INPUT_FILL


def _cond_type_fill(cond_type: str | None) -> PatternFill | None:
    if not cond_type:
        return None
    key = str(cond_type).strip().lower()
    if key == "hi":
        return HI_FILL
    if key == "lo":
        return LO_FILL
    if key == "force":
        return FORCE_FILL
    return None


def _style_header_row(ws, row: int, ncol: int) -> None:
    for col in range(1, ncol + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _cell_text_width(val) -> int:
    if val is None:
        return 0
    text = str(val)
    if "\n" in text:
        return max(len(line) for line in text.split("\n"))
    return len(text)


def _autosize_columns(ws, max_col: int, min_width: int = 9, max_width: int = 28) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value is not None:
                    best = max(best, min(_cell_text_width(cell.value) + 2, max_width))
        ws.column_dimensions[letter].width = best


def _collect_signal_list_entries() -> list[str]:
    nodes = [(row[0], row[1]) for row in NODE_ROWS]
    return build_signal_dropdown_entries(nodes, extra=["RSMRST_N"])


def _write_lists_sheet(wb: Workbook, pulses: list[str]) -> list[str]:
    ws = wb.create_sheet(SHEET_LISTS)
    ws.cell(row=1, column=1, value="Signal names (incl. presets)").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Pulse names (synced from Config)").font = Font(bold=True)
    entries = _collect_signal_list_entries()
    for i, name in enumerate(entries, start=2):
        ws.cell(row=i, column=1, value=name)
    for i, name in enumerate(pulses, start=2):
        ws.cell(row=i, column=2, value=name)
    sig_last = len(entries) + 1
    pulse_last = max(len(pulses) + 1, 2)
    wb.defined_names.add(
        DefinedName("SignalList", attr_text=f"{SHEET_LISTS}!$A$2:$A${sig_last}")
    )
    wb.defined_names.add(
        DefinedName("PulseList", attr_text=f"{SHEET_LISTS}!$B$2:$B${pulse_last}")
    )
    ws.sheet_state = "hidden"
    return entries


def _add_validations(ws_nodes, ws_cond, ws_input) -> None:
    type_dv = DataValidation(type="list", formula1='"Input,Output"', allow_blank=True)
    bin_dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)

    for dv in (type_dv, bin_dv):
        dv.error = "Pick from the list or see Instructions"
        dv.errorTitle = "Invalid value"

    nrows = 200
    ws_nodes.add_data_validation(type_dv)
    type_dv.add(f"B{DATA_START_ROW}:B{nrows}")
    ws_nodes.add_data_validation(bin_dv)
    for c, (key, _, _) in enumerate(_node_headers(), start=1):
        if key in ("init", "force_val"):
            letter = get_column_letter(c)
            bin_dv.add(f"{letter}{DATA_START_ROW}:{letter}{nrows}")
    pulse_dv = DataValidation(type="list", formula1="=PulseList", allow_blank=True)
    pulse_dv.error = "Pick from the list or see Instructions"
    pulse_dv.errorTitle = "Invalid value"
    ws_nodes.add_data_validation(pulse_dv)
    pulse_dv.add(f"H{DATA_START_ROW}:J{nrows}")

    cond_type_dv = DataValidation(type="list", formula1='"Hi,Lo,Force"', allow_blank=True)
    operation_dv = DataValidation(type="list", formula1='"AND,OR,XOR"', allow_blank=True)
    group_inv_out_dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    signal_out_dv = DataValidation(type="list", formula1="=SignalList", allow_blank=True)
    for dv in (cond_type_dv, operation_dv, group_inv_out_dv, signal_out_dv):
        dv.error = "Pick from the list or see Instructions"
        dv.errorTitle = "Invalid value"

    ws_cond.add_data_validation(cond_type_dv)
    ws_cond.add_data_validation(operation_dv)
    ws_cond.add_data_validation(group_inv_out_dv)
    ws_cond.add_data_validation(signal_out_dv)
    cond_type_dv.add(f"B{DATA_START_ROW}:B{nrows}")
    operation_dv.add(f"C{DATA_START_ROW}:C{nrows}")
    group_inv_out_dv.add(f"D{DATA_START_ROW}:D{nrows}")
    sig_start = get_column_letter(COND_META_COLS + 1)
    sig_end = get_column_letter(_cond_signal_end_col())
    signal_out_dv.add(f"{sig_start}{DATA_START_ROW}:{sig_end}{nrows}")

    side_dv = DataValidation(type="list", formula1='"Hi,Lo"', allow_blank=True)
    mode_dv = DataValidation(
        type="list",
        formula1='"Low (0),High (1),Custom wave,Signal cond."',
        allow_blank=True,
    )
    group_inv_in_dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    operation_in_dv = DataValidation(type="list", formula1='"AND,OR,XOR"', allow_blank=True)
    signal_in_dv = DataValidation(type="list", formula1="=SignalList", allow_blank=True)
    for dv in (side_dv, mode_dv, operation_in_dv, group_inv_in_dv, signal_in_dv):
        dv.error = "Pick from the list or see Instructions"
        dv.errorTitle = "Invalid value"

    ws_input.add_data_validation(side_dv)
    ws_input.add_data_validation(mode_dv)
    ws_input.add_data_validation(operation_in_dv)
    ws_input.add_data_validation(group_inv_in_dv)
    ws_input.add_data_validation(signal_in_dv)
    side_dv.add(f"B{DATA_START_ROW}:B{nrows}")
    mode_dv.add(f"C{DATA_START_ROW}:C{nrows}")
    operation_in_dv.add(f"E{DATA_START_ROW}:E{nrows}")
    group_inv_in_dv.add(f"F{DATA_START_ROW}:F{nrows}")
    in_sig_start = get_column_letter(INPUT_META_COLS + 1)
    in_sig_end = get_column_letter(_input_cond_signal_end_col())
    signal_in_dv.add(f"{in_sig_start}{DATA_START_ROW}:{in_sig_end}{nrows}")


def _write_config_sheet(ws) -> str:
    pulses_csv = "Pulse_1us,Pulse_2ms"
    for r, row in enumerate(CONFIG_ROWS, start=1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
        if row and row[0] == "pulses" and len(row) > 3:
            pulses_csv = str(row[3])
    _style_header_row(ws, 1, 4)
    _autosize_columns(ws, 4)
    for r in range(1, 50):
        if str(ws.cell(row=r, column=1).value or "").strip().lower() == "pulses":
            return str(ws.cell(row=r, column=4).value or pulses_csv)
    return pulses_csv


def _write_nodes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_NODES)
    headers = _node_headers()
    apply_nodes_sheet_header_rows(ws)

    for i, row_data in enumerate(NODE_ROWS):
        fill = row_data[-1]
        values = row_data[:-1]
        r = DATA_START_ROW + i
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val if val is not None else None)
            cell.fill = fill

    wb.defined_names.add(
        DefinedName(
            "NodeNames",
            attr_text=f"'{SHEET_NODES}'!$A${DATA_START_ROW}:$A$200",
        )
    )
    ws.freeze_panes = f"A{DATA_START_ROW}"


def _write_cond_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_COND)
    headers = _cond_headers()
    total_cols = _cond_total_cols()
    sig_end_col = _cond_signal_end_col()

    for c in range(total_cols + 1, sig_end_col + 1):
        ws.cell(row=1, column=c, value=f"signal{c - COND_META_COLS}")

    for c, (key, label, hint) in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=key)
        ws.cell(row=2, column=c, value=label)
        ws.cell(row=3, column=c, value=hint)
    if COND_SIGNAL_INIT_COLS > 1:
        ws.merge_cells(
            start_row=2,
            start_column=COND_META_COLS + 1,
            end_row=2,
            end_column=total_cols,
        )
        ws.cell(row=2, column=COND_META_COLS + 1).alignment = Alignment(horizontal="center")

    _style_header_row(ws, 1, sig_end_col)
    for c in range(1, sig_end_col + 1):
        ws.cell(row=2, column=c).font = Font(bold=True)
        ws.cell(row=3, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    block_start = DATA_START_ROW
    for i, row_data in enumerate(COND_ROWS):
        r = DATA_START_ROW + i
        output_name = row_data[0] if row_data else ""
        cond_type = row_data[1] if len(row_data) > 1 else None
        operation = row_data[2] if len(row_data) > 2 else "AND"
        group_inv = row_data[3] if len(row_data) > 3 else None
        signals = row_data[4:] if len(row_data) > 4 else ()

        if output_name:
            block_start = r
        ws.cell(row=r, column=1, value=output_name if output_name else None)
        ws.cell(row=r, column=2, value=cond_type)
        ws.cell(row=r, column=3, value=operation)
        ws.cell(row=r, column=4, value=group_inv)
        for si, sig in enumerate(signals, start=1):
            if sig is not None:
                ws.cell(row=r, column=COND_META_COLS + si, value=sig)

        fill = _cond_type_fill(cond_type)
        if fill:
            for c in range(1, sig_end_col + 1):
                ws.cell(row=r, column=c).fill = fill

    block_end = DATA_START_ROW + len(COND_ROWS) - 1
    if block_end > block_start:
        ws.merge_cells(
            start_row=block_start,
            start_column=1,
            end_row=block_end,
            end_column=1,
        )

    ws.freeze_panes = f"A{DATA_START_ROW}"
    _autosize_columns(ws, total_cols, max_width=16)


def _write_input_cond_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_INPUT_COND)
    headers = _input_cond_headers()
    total_cols = _input_cond_total_cols()
    sig_end_col = _input_cond_signal_end_col()

    for c in range(total_cols + 1, sig_end_col + 1):
        ws.cell(row=1, column=c, value=f"signal{c - INPUT_META_COLS}")

    for c, (key, label, hint) in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=key)
        ws.cell(row=2, column=c, value=label)
        ws.cell(row=3, column=c, value=hint)
    if INPUT_COND_SIGNAL_INIT_COLS > 1:
        ws.merge_cells(
            start_row=2,
            start_column=INPUT_META_COLS + 1,
            end_row=2,
            end_column=total_cols,
        )
        ws.cell(row=2, column=INPUT_META_COLS + 1).alignment = Alignment(horizontal="center")

    _style_header_row(ws, 1, sig_end_col)
    for c in range(1, sig_end_col + 1):
        ws.cell(row=2, column=c).font = Font(bold=True)
        ws.cell(row=3, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    block_start = DATA_START_ROW
    for i, row_data in enumerate(INPUT_COND_ROWS):
        r = DATA_START_ROW + i
        input_name = row_data[0] if row_data else ""
        side = row_data[1] if len(row_data) > 1 else None
        mode = row_data[2] if len(row_data) > 2 else None
        wave = row_data[3] if len(row_data) > 3 else None
        operation = row_data[4] if len(row_data) > 4 else "AND"
        group_inv = row_data[5] if len(row_data) > 5 else None
        signals = row_data[6:] if len(row_data) > 6 else ()

        if input_name:
            block_start = r
        ws.cell(row=r, column=1, value=input_name if input_name else None)
        ws.cell(row=r, column=2, value=side)
        ws.cell(row=r, column=3, value=mode)
        ws.cell(row=r, column=4, value=wave if wave is not None else None)
        ws.cell(row=r, column=5, value=operation)
        ws.cell(row=r, column=6, value=group_inv)
        for si, sig in enumerate(signals, start=1):
            if sig is not None:
                ws.cell(row=r, column=INPUT_META_COLS + si, value=sig)

        fill = _input_side_fill(side)
        if fill:
            for c in range(1, sig_end_col + 1):
                ws.cell(row=r, column=c).fill = fill

    block_end = DATA_START_ROW + len(INPUT_COND_ROWS) - 1
    if block_end > block_start:
        ws.merge_cells(
            start_row=block_start,
            start_column=1,
            end_row=block_end,
            end_column=1,
        )

    ws.freeze_panes = f"A{DATA_START_ROW}"
    _autosize_columns(ws, total_cols, max_width=16)


def _write_legend_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet(SHEET_LEGEND)
    ws.cell(row=1, column=1, value="Sheet").font = Font(bold=True)
    ws.cell(row=1, column=2, value="key").font = Font(bold=True)
    ws.cell(row=1, column=3, value="Label").font = Font(bold=True)
    ws.cell(row=1, column=4, value="Description").font = Font(bold=True)
    _style_header_row(ws, 1, 4)
    row = 2
    for key, label, hint in _node_headers():
        ws.cell(row=row, column=1, value=SHEET_NODES)
        ws.cell(row=row, column=2, value=key)
        ws.cell(row=row, column=3, value=label)
        ws.cell(row=row, column=4, value=hint)
        row += 1
    for key, label, hint in _cond_headers():
        ws.cell(row=row, column=1, value=SHEET_COND)
        ws.cell(row=row, column=2, value=key)
        ws.cell(row=row, column=3, value=label)
        ws.cell(row=row, column=4, value=hint)
        row += 1
    for key, label, hint in _input_cond_headers():
        ws.cell(row=row, column=1, value=SHEET_INPUT_COND)
        ws.cell(row=row, column=2, value=key)
        ws.cell(row=row, column=3, value=label)
        ws.cell(row=row, column=4, value=hint)
        row += 1
    _autosize_columns(ws, 4, max_width=48)


def _write_instructions_sheet(ws) -> None:
    for r, line in enumerate(INSTRUCTIONS, start=1):
        cell = ws.cell(row=r, column=1, value=line)
        if r == 1:
            cell.font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 90


def _sync_cond_output_names() -> None:
    """Ensure COND_ROWS includes 1 Hi + 1 Lo + 1 Force per Output (template generation)."""
    outputs = [r[0] for r in NODE_ROWS if r[1] == "Output"]
    blocks: dict[str, list[tuple]] = {}
    current = ""
    for row in COND_ROWS:
        name = row[0] or current
        if row[0]:
            current = row[0]
        if name:
            blocks.setdefault(name, []).append(row)
    for name in outputs:
        if name not in blocks:
            blocks[name] = [
                (name, "Hi", "AND", "N"),
                ("", "Lo", "AND", "N"),
                ("", "Force", "AND", "N"),
            ]
    COND_ROWS.clear()
    for name in outputs:
        COND_ROWS.extend(blocks[name])


def _sync_input_cond_names() -> None:
    """Ensure INPUT_COND_ROWS includes 1 Hi + 1 Lo per Input (template generation)."""
    inputs = [r[0] for r in NODE_ROWS if r[1] == "Input"]
    blocks: dict[str, list[tuple]] = {}
    current = ""
    for row in INPUT_COND_ROWS:
        name = row[0] or current
        if row[0]:
            current = row[0]
        if name:
            blocks.setdefault(name, []).append(row)
    for name in inputs:
        if name not in blocks:
            blocks[name] = [
                (name, "Hi", "Signal cond.", None, "AND", "N"),
                ("", "Lo", "Low (0)", None, "AND", "N"),
            ]
    INPUT_COND_ROWS.clear()
    for name in inputs:
        INPUT_COND_ROWS.extend(blocks[name])


VBA_BIN = os.path.join(ROOT, "templates", "vba", "vbaProject.bin")


def _extract_vba_bin(xlsm_path: str, bin_path: str) -> None:
    import zipfile

    with zipfile.ZipFile(xlsm_path, "r") as zf:
        data = zf.read("xl/vbaProject.bin")
    os.makedirs(os.path.dirname(bin_path), exist_ok=True)
    with open(bin_path, "wb") as fh:
        fh.write(data)


def _embed_vba_macros(xlsx_path: str, xlsm_path: str) -> None:
    scripts_dir = os.path.dirname(os.path.abspath(__file__))
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)

    from embed_excel_vba import add_sync_button_inplace, embed_vba, refresh_vba_inplace

    try:
        embed_vba(xlsx_path, xlsm_path)
        _extract_vba_bin(xlsm_path, VBA_BIN)
        return
    except Exception as exc:
        print(f"COM embed failed ({exc}); trying zip inject...")

    if os.path.isfile(VBA_BIN):
        from inject_vba_zip import inject_vba_zip

        inject_vba_zip(xlsx_path, xlsm_path, VBA_BIN)
        try:
            refresh_vba_inplace(xlsm_path)
            _extract_vba_bin(xlsm_path, VBA_BIN)
            return
        except Exception as exc:
            print(f"VBA refresh failed ({exc}); falling back to zip-only inject")
    else:
        print("No vbaProject.bin; cannot zip-inject without prior COM embed")
        return

    try:
        add_sync_button_inplace(xlsm_path)
    except Exception as exc:
        print(f"Sync button failed ({exc}); run RequestSyncFromNodes from Alt+F8")


def main() -> None:
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    _sync_cond_output_names()
    _sync_input_cond_names()
    wb = Workbook()
    ws_config = wb.active
    ws_config.title = "Config"
    pulses_csv = _write_config_sheet(ws_config)
    pulses = [p.strip() for p in pulses_csv.split(",") if p.strip()]
    ws_inst = wb.create_sheet(SHEET_INSTRUCTIONS, 0)
    _write_instructions_sheet(ws_inst)
    _write_lists_sheet(wb, pulses)
    _write_nodes_sheet(wb)
    _write_cond_sheet(wb)
    _write_input_cond_sheet(wb)
    _write_legend_sheet(wb)
    ws_nodes = wb[SHEET_NODES]
    ws_cond = wb[SHEET_COND]
    ws_input = wb[SHEET_INPUT_COND]
    _add_validations(ws_nodes, ws_cond, ws_input)
    wb.save(TMP_XLSX)
    try:
        _embed_vba_macros(TMP_XLSX, OUT_PATH)
        print(f"Wrote: {OUT_PATH}")
    except Exception as exc:
        fallback = os.path.join(ROOT, "templates", "powerseq_nodes_template.xlsx")
        import shutil

        shutil.copy2(TMP_XLSX, fallback)
        print(f"VBA 嵌入失敗（{exc}）")
        print(f"已改輸出無巨集版: {fallback}")
        print("若需 xlsm：安裝 Excel + pywin32，並啟用「信任對 VBA 專案物件模型的存取」後重跑。")
    finally:
        if os.path.isfile(TMP_XLSX):
            try:
                os.remove(TMP_XLSX)
            except OSError:
                pass


if __name__ == "__main__":
    main()
