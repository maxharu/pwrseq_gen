"""Tests for excel_export.py round-trip."""
import os
import sys
import tempfile
import zipfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src"))

from dataclasses import replace

from config_models import PowerRail, PowerSeqConfig
from excel_export import export_powerseq_to_excel, format_signal_cell
from excel_import import load_powerseq_from_excel
from timing_sim import DEP_LOW, TimingScenario

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_XLSX = os.path.join(ROOT, "templates", "powerseq_nodes_template.xlsx")
GOLDEN_XLSM = os.path.join(ROOT, "templates", "powerseq_nodes_template_new.xlsm")
HAS_EXPORT_TEMPLATE = os.path.isfile(TEMPLATE_XLSX) or os.path.isfile(GOLDEN_XLSM)


def _load_demo_config() -> PowerSeqConfig:
    if os.path.isfile(TEMPLATE_XLSX):
        return load_powerseq_from_excel(TEMPLATE_XLSX)
    return load_powerseq_from_excel(GOLDEN_XLSM)


class TestFormatSignalCell:
    def test_presets_and_suffix(self):
        assert format_signal_cell(DEP_LOW, False, "self") == "Low"
        assert format_signal_cell("OUT", False, "hi") == "OUT|Hi Cond"
        assert format_signal_cell("SIG", True, "self") == "!SIG"


@pytest.mark.skipif(not HAS_EXPORT_TEMPLATE, reason="excel template missing")
class TestExportRoundtrip:
    def test_export_then_import(self):
        cfg = _load_demo_config()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            export_powerseq_to_excel(cfg, out)
            cfg2 = load_powerseq_from_excel(out)
            assert cfg2.module_name == cfg.module_name
            assert [r.name for r in cfg2.rails] == [r.name for r in cfg.rails]
            for src, dst in zip(cfg.rails, cfg2.rails):
                if src.seq_type == "output":
                    assert dst.get_hi_groups() == src.get_hi_groups()
                    assert dst.get_lo_groups() == src.get_lo_groups()
                if src.seq_type == "input" and src.hi_mode != "depends":
                    assert dst.hi_mode == src.hi_mode
        finally:
            os.unlink(out)

    def test_export_minimal_config(self):
        cfg = PowerSeqConfig(
            module_name="TEST",
            pulses=["Pulse_1us"],
            rails=[
                PowerRail("A", seq_type="input", deb_cycle_hi=1, deb_cycle_lo=1),
                PowerRail(
                    "B",
                    seq_type="output",
                    cycle_hi=3,
                    depends_on_hi_groups=[["A"]],
                ),
            ],
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            export_powerseq_to_excel(cfg, out)
            cfg2 = load_powerseq_from_excel(out)
            assert cfg2.module_name == "TEST"
            assert len(cfg2.rails) == 2
            b = cfg2.rails[1]
            assert b.get_hi_groups() == [["A"]]
        finally:
            os.unlink(out)

    def test_export_preserves_nodes_row3_multiline_hints(self):
        cfg = _load_demo_config()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            export_powerseq_to_excel(cfg, out)
            from openpyxl import load_workbook

            ws = load_workbook(out)["Nodes"]
            assert "\n" in str(ws.cell(3, 1).value)
            assert "\n" in str(ws.cell(3, 3).value)
            assert ws.cell(3, 1).alignment.wrap_text
        finally:
            os.unlink(out)

    def test_export_xlsm_is_valid_zip(self):
        cfg = _load_demo_config()
        with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as f:
            out = f.name
        try:
            export_powerseq_to_excel(cfg, out)
            assert zipfile.is_zipfile(out)
            cfg2 = load_powerseq_from_excel(out)
            assert len(cfg2.rails) == len(cfg.rails)
        finally:
            os.unlink(out)

    @pytest.mark.skipif(not os.path.isfile(GOLDEN_XLSM), reason="golden xlsm missing")
    def test_export_xlsm_has_sync_button(self):
        cfg = _load_demo_config()
        with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as f:
            out = f.name
        try:
            export_powerseq_to_excel(cfg, out)
            with zipfile.ZipFile(out) as zf:
                nodes_part = None
                import xml.etree.ElementTree as ET

                wb = ET.fromstring(zf.read("xl/workbook.xml"))
                m = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
                rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
                rns = "http://schemas.openxmlformats.org/package/2006/relationships"
                rel_map = {
                    el.get("Id"): el.get("Target")
                    for el in rels.findall(f"{{{rns}}}Relationship")
                }
                for sh in wb.findall(f".//{{{m}}}sheet"):
                    if sh.get("name") == "Nodes":
                        rid = sh.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                        target = rel_map.get(rid)
                        nodes_part = f"xl/{target}"
                        break
                assert nodes_part is not None
                sheet_xml = zf.read(nodes_part).decode("utf-8")
                assert "btnSyncOutput" in sheet_xml
                assert "RequestSyncFromNodes" in sheet_xml
                assert "xl/drawings/drawing1.xml" in zf.namelist()
        finally:
            os.unlink(out)

    def test_export_with_updated_timing_scenario(self):
        cfg = _load_demo_config()
        cfg2 = replace(cfg, timing_scenario={"steps": 77})
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            export_powerseq_to_excel(cfg2, out)
            loaded = load_powerseq_from_excel(out)
            assert loaded.timing_scenario is not None
            assert loaded.timing_scenario["steps"] == 77
        finally:
            os.unlink(out)

    def test_export_clears_rows_beyond_shrunk_nodes(self):
        """Nodes 從 250 筆縮到 230 筆時，舊資料列應被清除。"""
        from openpyxl import load_workbook

        large = PowerSeqConfig(
            module_name="SHRINK",
            pulses=["Pulse_1us"],
            rails=[
                PowerRail(
                    f"OUT{i}",
                    seq_type="output",
                    cycle_hi=1,
                    depends_on_hi_groups=[["IN_A"]],
                )
                for i in range(250)
            ]
            + [
                PowerRail("IN_A", seq_type="input", deb_cycle_hi=1, deb_cycle_lo=1),
            ],
        )
        small = replace(
            large,
            rails=large.rails[:230] + [large.rails[-1]],
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            out = f.name
        try:
            export_powerseq_to_excel(large, out)
            wb = load_workbook(out)
            ws = wb["Nodes"]
            assert ws.cell(4 + 249, 1).value == "OUT249"
            wb.close()

            export_powerseq_to_excel(small, out)
            wb = load_workbook(out)
            ws = wb["Nodes"]
            assert ws.cell(4 + 229, 1).value == "OUT229"
            assert ws.cell(4 + 230, 1).value == "IN_A"
            assert ws.cell(4 + 231, 1).value is None
            assert ws.cell(4 + 249, 1).value is None
            wb.close()
        finally:
            os.unlink(out)
