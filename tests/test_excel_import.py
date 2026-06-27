"""Tests for excel_import.py"""
import os
import sys
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src"))

from config_models import PowerRail, PowerSeqConfig
from excel_import import (
    _parse_input_cond_meta,
    _parse_output_cond_meta,
    build_signal_dropdown_entries,
    load_powerseq_from_excel,
    parse_signal_cell,
)
from timing_sim import DEP_HIGH, DEP_LOW

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_XLSX = os.path.join(ROOT, "templates", "powerseq_nodes_template.xlsx")
HAS_TEMPLATE = os.path.isfile(TEMPLATE_XLSX)


class TestParseCondMeta:
    def test_output_operation_and_group_inv(self):
        row = ("OUT", "Hi", "XOR", "Y") + (None,) * 10
        assert _parse_output_cond_meta(row) == ("xor", True)

    def test_output_legacy_group_inv_col3(self):
        row = ("OUT", "Hi", "Y") + (None,) * 10
        assert _parse_output_cond_meta(row) == ("and", True)

    def test_input_operation_and_group_inv(self):
        row = ("IN", "Hi", "Signal cond.", None, "OR", "N") + (None,) * 10
        assert _parse_input_cond_meta(row) == ("or", False)

    def test_input_legacy_group_inv_col5(self):
        row = ("IN", "Hi", "Signal cond.", None, "Y") + (None,) * 10
        assert _parse_input_cond_meta(row) == ("and", True)


class TestParseSignalCell:
    def test_invert_and_use_suffix(self):
        assert parse_signal_cell("!RSMRST_N") == ("RSMRST_N", True, "self")
        assert parse_signal_cell("OUT|Hi Cond") == ("OUT", False, "hi")
        assert parse_signal_cell("Low") == (DEP_LOW, False, "self")
        assert parse_signal_cell("High") == (DEP_HIGH, False, "self")


class TestBuildSignalDropdownEntries:
    def test_includes_inverted_node_names(self):
        nodes = [("IN_A", "input"), ("OUT_B", "output")]
        entries = build_signal_dropdown_entries(nodes)
        assert entries[:2] == ["High", "Low"]
        assert "IN_A" in entries
        assert "!IN_A" in entries
        assert "OUT_B" in entries
        assert "!OUT_B" in entries
        assert "OUT_B|Hi Cond" in entries
        assert "!OUT_B|Hi Cond" not in entries


class TestLastUsedRow:
    def test_scan_finds_last_row_beyond_template_floor(self):
        from openpyxl import Workbook

        from excel_import import DATA_START_ROW, last_used_row

        wb = Workbook()
        ws = wb.active
        ws.cell(DATA_START_ROW, 1, value="FIRST")
        ws.cell(350, 1, value="LAST")
        assert last_used_row(ws, min_row=DATA_START_ROW, max_col=5) == 350

    def test_scan_empty_returns_below_start(self):
        from openpyxl import Workbook

        from excel_import import DATA_START_ROW, last_used_row

        wb = Workbook()
        ws = wb.active
        assert last_used_row(ws, min_row=DATA_START_ROW, max_col=5) == DATA_START_ROW - 1


@pytest.mark.skipif(not HAS_TEMPLATE, reason="excel template missing")
class TestLoadTimingConfigGlobals:
    def test_timing_steps_hscale_from_config_sheet(self):
        from excel_export import export_powerseq_to_excel

        cfg = PowerSeqConfig(
            module_name="WD_CFG",
            pulses=["Pulse_1us"],
            rails=[
                PowerRail(
                    "OUT_ONLY",
                    seq_type="output",
                    cycle_hi=1,
                    cycle_lo=1,
                    depends_on_hi=["__HIGH__"],
                ),
            ],
            timing_scenario={"steps": 120, "hscale": 3},
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            export_powerseq_to_excel(cfg, path)
            loaded = load_powerseq_from_excel(path)
            assert loaded.timing_scenario is not None
            assert loaded.timing_scenario["steps"] == 120
            assert loaded.timing_scenario["hscale"] == 3
        finally:
            os.unlink(path)

    def test_timing_steps_without_input_conditions(self):
        from openpyxl import load_workbook

        from excel_export import export_powerseq_to_excel

        cfg = PowerSeqConfig(
            module_name="WD_NO_IN",
            pulses=["Pulse_1us"],
            rails=[
                PowerRail(
                    "OUT_ONLY",
                    seq_type="output",
                    cycle_hi=1,
                    cycle_lo=1,
                    depends_on_hi=["__HIGH__"],
                ),
            ],
            timing_scenario={"steps": 88},
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            export_powerseq_to_excel(cfg, path)
            wb = load_workbook(path)
            if "Input Conditions" in wb.sheetnames:
                del wb["Input Conditions"]
            wb.save(path)
            wb.close()
            loaded = load_powerseq_from_excel(path)
            assert loaded.timing_scenario is not None
            assert loaded.timing_scenario["steps"] == 88
        finally:
            os.unlink(path)
