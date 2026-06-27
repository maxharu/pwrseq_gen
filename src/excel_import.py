"""
Load PowerSeqConfig from Excel node template (.xlsx / .xlsm).

Sheet detection uses row-1 column-A key (same as VBA SetSheetRefs):
  key, name, output_name, input_name
"""
from __future__ import annotations

import os
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise ImportError("Excel import requires openpyxl: pip install openpyxl") from e

from config_models import DEFAULT_PULSE, PowerRail, PowerSeqConfig, normalize_pulse_name
from group_logic import is_legacy_group_inv_cell, parse_intra_op_cell
from timing_sim import DEP_HIGH, DEP_LOW

DATA_START_ROW = 4
MAX_DATA_ROW_CAP = 5000
COND_META_COLS = 4
INPUT_META_COLS = 6  # input_name, cond_side, mode, wave, operation, group_inv
COND_SIGNAL_MAX_COLS = 49
INPUT_COND_SIGNAL_MAX_COLS = 49
NODES_MAX_COL = 10
OUTPUT_COND_MAX_COL = COND_META_COLS + COND_SIGNAL_MAX_COLS
INPUT_COND_MAX_COL = INPUT_META_COLS + INPUT_COND_SIGNAL_MAX_COLS
CONFIG_MAX_ROW = 59
CONFIG_MAX_COL = 4

SIGNAL_PRESETS_HIGH = frozenset({"High", "固定為高", "Constant High"})
SIGNAL_PRESETS_LOW = frozenset({"Low", "固定為低", "Constant Low"})
OUTPUT_USE_SUFFIXES = ("|Hi Cond", "|Lo Cond", "|Force Cond")


def build_signal_dropdown_entries(
    nodes: list[tuple[str, str]],
    *,
    extra: list[str] | None = None,
) -> list[str]:
    """Build Lists column-A entries: presets, each node, !node, output use suffixes, extras."""
    entries = ["High", "Low"]
    seen = set(entries)
    for name, typ in nodes:
        if not name:
            continue
        for item in (name, f"!{name}"):
            if item not in seen:
                entries.append(item)
                seen.add(item)
        if str(typ).strip().lower() == "output":
            for suffix in OUTPUT_USE_SUFFIXES:
                item = f"{name}{suffix}"
                if item not in seen:
                    entries.append(item)
                    seen.add(item)
    for sig in sorted(extra or ()):
        if sig and sig not in seen:
            entries.append(sig)
            seen.add(sig)
    return entries


INPUT_MODE_MAP = {
    "low (0)": "constant_0",
    "high (1)": "constant_1",
    "custom wave": "custom",
    "signal cond.": "depends",
    "signal cond": "depends",
}


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _cell_int(val: Any, default: int) -> int:
    s = _cell_str(val)
    if not s:
        return default
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return default


def _is_yes(val: Any) -> bool:
    s = _cell_str(val).lower()
    return s in ("y", "yes", "1", "true")


def _row_val(row: tuple, col_1based: int, default: Any = None) -> Any:
    idx = col_1based - 1
    if idx < 0 or idx >= len(row):
        return default
    return row[idx]


def _row_has_value(row: tuple) -> bool:
    return any(v is not None and str(v).strip() != "" for v in row)


def last_used_row(ws, *, min_row: int, max_col: int) -> int:
    """掃描 cols 1..max_col，回傳 min_row 起最後一列有內容的列號（1-based）；無則 min_row-1。"""
    last = min_row - 1
    for i, row in enumerate(
        ws.iter_rows(
            min_row=min_row,
            max_row=MAX_DATA_ROW_CAP,
            max_col=max_col,
            values_only=True,
        )
    ):
        if _row_has_value(row):
            last = min_row + i
    return last


def data_area_end_row(ws, *, new_row_count: int = 0, max_col: int) -> int:
    """Import 掃描 / export 清除的最後列：max(掃描結果, 本次寫入)，上限 MAX_DATA_ROW_CAP。"""
    used_end = last_used_row(ws, min_row=DATA_START_ROW, max_col=max_col)
    write_end = (
        DATA_START_ROW + new_row_count - 1 if new_row_count > 0 else DATA_START_ROW - 1
    )
    return min(max(used_end, write_end), MAX_DATA_ROW_CAP)


def _materialize_rows(ws, *, min_row: int, max_row: int, max_col: int) -> list[tuple]:
    """讀取固定矩形範圍（Config 等小表）。"""
    if max_row < min_row:
        return []
    return list(
        ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            max_col=max_col,
            values_only=True,
        )
    )


def _data_sheet_rows(ws, *, max_col: int) -> list[tuple]:
    """讀取資料區：掃描實際最後列。"""
    end = data_area_end_row(ws, new_row_count=0, max_col=max_col)
    return _materialize_rows(
        ws, min_row=DATA_START_ROW, max_row=end, max_col=max_col,
    )


def _sheet_rows(ws, *, min_row: int, max_row: int, max_col: int) -> list[tuple]:
    """讀取固定矩形範圍（Config 等小表）。"""
    return _materialize_rows(ws, min_row=min_row, max_row=max_row, max_col=max_col)


def _sheet_a1(ws) -> str:
    row = next(
        ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1, values_only=True),
        None,
    )
    if not row:
        return ""
    return _cell_str(row[0])


def _find_sheet(wb, header_key: str):
    key = header_key.lower()
    for name in wb.sheetnames:
        ws = wb[name]
        if _sheet_a1(ws).lower() == key:
            return ws
    raise ValueError(f"Sheet with header '{header_key}' not found in workbook")


def _config_lookup(rows: list[tuple]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for row in rows:
        k = _cell_str(_row_val(row, 1)).lower()
        if k:
            out[k] = _row_val(row, 4)
    return out


def _config_value(lookup: dict[str, Any], key: str, default: Any = None) -> Any:
    return lookup.get(key.lower(), default)


def _parse_pulses(raw: Any) -> list[str]:
    if raw is None:
        return [DEFAULT_PULSE]
    parts = [_cell_str(p) for p in str(raw).split(",")]
    parts = [p for p in parts if p]
    if not parts:
        return [DEFAULT_PULSE]
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        n = normalize_pulse_name(p)
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out or [DEFAULT_PULSE]


def parse_signal_cell(raw: Any) -> tuple[str, bool, str] | None:
    """Return (dep_name, inverted, use) or None if empty."""
    s = _cell_str(raw)
    if not s:
        return None
    inv = False
    if s.startswith("!"):
        inv = True
        s = s[1:].strip()
    if s in SIGNAL_PRESETS_HIGH:
        return (DEP_HIGH, False, "self")
    if s in SIGNAL_PRESETS_LOW:
        return (DEP_LOW, False, "self")
    use = "self"
    for suffix in OUTPUT_USE_SUFFIXES:
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
            if suffix == "|Hi Cond":
                use = "hi"
            elif suffix == "|Lo Cond":
                use = "lo"
            else:
                use = "force"
            break
    if not s:
        return None
    return (s, inv, use)


def _read_signal_cells_row(
    row: tuple, start_col_1based: int, max_cols: int,
) -> list[tuple[str, bool, str]]:
    signals: list[tuple[str, bool, str]] = []
    start = start_col_1based - 1
    end = min(len(row), start + max_cols)
    for c in range(start, end):
        parsed = parse_signal_cell(row[c])
        if parsed:
            signals.append(parsed)
    return signals


def _parse_input_cond_meta(row: tuple) -> tuple[str, bool]:
    """Return (intra_op, group_inv) from cols 5–6 (legacy: col5 = group inv only)."""
    c5 = _row_val(row, 5)
    if is_legacy_group_inv_cell(c5):
        return "and", _is_yes(c5)
    return parse_intra_op_cell(c5), _is_yes(_row_val(row, 6))


def _parse_output_cond_meta(row: tuple) -> tuple[str, bool]:
    """Return (intra_op, group_inv) from row cols 3–4 (legacy: col3 = group inv only)."""
    c3 = _row_val(row, 3)
    if is_legacy_group_inv_cell(c3):
        return "and", _is_yes(c3)
    return parse_intra_op_cell(c3), _is_yes(_row_val(row, 4))


def _apply_condition_groups(
    rail: PowerRail,
    kind: str,
    row_groups: list[tuple[list[tuple[str, bool, str]], bool, str]],
) -> None:
    names_g: list[list[str]] = []
    inv_g: list[list[bool]] = []
    use_g: list[list[str]] = []
    ginv: list[bool] = []
    intra: list[str] = []
    for signals, group_inv, intra_op in row_groups:
        if not signals:
            continue
        names_g.append([s[0] for s in signals])
        inv_g.append([s[1] for s in signals])
        use_g.append([s[2] for s in signals])
        ginv.append(group_inv)
        intra.append(intra_op)

    prefix = f"depends_on_{kind}"
    setattr(rail, f"{prefix}_groups", names_g)
    setattr(rail, f"{prefix}_inv_groups", inv_g)
    setattr(rail, f"{prefix}_use_groups", use_g)
    setattr(rail, f"{prefix}_group_inv", ginv)
    setattr(rail, f"{prefix}_intra_op", intra)

    flat = [n for g in names_g for n in g]
    setattr(rail, prefix, flat)

    inv_dict: dict[str, bool] = {}
    use_dict: dict[str, str] = {}
    for gi, group in enumerate(names_g):
        for ii, name in enumerate(group):
            if name not in inv_dict:
                inv_dict[name] = inv_g[gi][ii]
                use_dict[name] = use_g[gi][ii]
    setattr(rail, f"{prefix}_inv", inv_dict)
    setattr(rail, f"{prefix}_use", use_dict)


def _timing_scenario_from_config_defaults(defaults: dict[str, Any]) -> dict:
    """Global timing steps/hscale from Config sheet (GUI toolbar reads this)."""
    out: dict = {"steps": defaults["timing_steps"]}
    if defaults["timing_hscale"] != 1:
        out["hscale"] = defaults["timing_hscale"]
    return out


def _parse_config_defaults(lookup: dict[str, Any]) -> dict[str, Any]:
    return {
        "module_name": _cell_str(_config_value(lookup, "module_name")) or "PWRSEQ_TOP",
        "pulses": _parse_pulses(_config_value(lookup, "pulses")),
        "default_cycle_hi": _cell_int(_config_value(lookup, "default_cycle_hi"), 8),
        "default_cycle_lo": _cell_int(_config_value(lookup, "default_cycle_lo"), 4),
        "default_cycle_force": _cell_int(_config_value(lookup, "default_cycle_force"), 2),
        "default_init": _cell_int(_config_value(lookup, "default_init"), 0),
        "default_force_val": _cell_int(_config_value(lookup, "default_force_val"), 0),
        "default_pulse": normalize_pulse_name(
            _cell_str(_config_value(lookup, "default_pulse")) or DEFAULT_PULSE
        ),
        "timing_steps": _cell_int(_config_value(lookup, "timing_steps"), 50),
        "timing_hscale": _cell_int(_config_value(lookup, "timing_hscale"), 1),
    }


def _parse_nodes(rows: list[tuple], defaults: dict[str, Any]) -> list[PowerRail]:
    rails: list[PowerRail] = []
    dp = defaults["default_pulse"]
    for row in rows:
        name = _cell_str(_row_val(row, 1))
        if not name:
            if rails:
                break
            continue
        seq_raw = _cell_str(_row_val(row, 2)).lower()
        seq_type = "input" if seq_raw == "input" else "output"

        cycle_hi = _cell_int(_row_val(row, 3), defaults["default_cycle_hi"])
        cycle_lo = _cell_int(_row_val(row, 4), defaults["default_cycle_lo"])
        cycle_force = _cell_int(_row_val(row, 5), defaults["default_cycle_force"])
        init = _cell_int(_row_val(row, 6), defaults["default_init"])
        force_val = _cell_int(_row_val(row, 7), defaults["default_force_val"])
        pulse_hi = normalize_pulse_name(_cell_str(_row_val(row, 8)) or dp)
        pulse_lo = normalize_pulse_name(_cell_str(_row_val(row, 9)) or dp)
        pulse_timing = normalize_pulse_name(_cell_str(_row_val(row, 10)) or dp)

        if seq_type == "input":
            deb_enable = cycle_hi != 0 or cycle_lo != 0
            rails.append(
                PowerRail(
                    name=name,
                    seq_type="input",
                    deb_enable=deb_enable,
                    deb_init=init,
                    deb_cycle_hi=cycle_hi,
                    deb_cycle_lo=cycle_lo,
                    deb_pulse=pulse_timing,
                    pulse_hi=pulse_hi,
                    pulse_lo=pulse_lo,
                )
            )
        else:
            rails.append(
                PowerRail(
                    name=name,
                    seq_type="output",
                    cycle_hi=cycle_hi,
                    cycle_lo=cycle_lo,
                    cycle_force=cycle_force,
                    init=init,
                    force_val=force_val,
                    pulse_hi=pulse_hi,
                    pulse_lo=pulse_lo,
                    pulse_force=pulse_timing,
                )
            )
    return rails


def _parse_output_conditions(rows: list[tuple], rails_by_name: dict[str, PowerRail]) -> None:
    current_output = ""
    pending: dict[str, list[tuple[list[tuple[str, bool, str]], bool, str]]] = {
        "hi": [],
        "lo": [],
        "force": [],
    }

    def flush(output_name: str) -> None:
        if not output_name:
            return
        rail = rails_by_name.get(output_name)
        if rail is None:
            return
        for kind in ("hi", "lo", "force"):
            groups = pending[kind]
            if groups:
                _apply_condition_groups(rail, kind, groups)
        pending["hi"] = []
        pending["lo"] = []
        pending["force"] = []

    for idx, row in enumerate(rows):
        row_num = DATA_START_ROW + idx
        name_cell = _cell_str(_row_val(row, 1))
        cond_type = _cell_str(_row_val(row, 2))
        if name_cell:
            if current_output:
                flush(current_output)
            current_output = name_cell
        if not current_output and not cond_type:
            if row_num > DATA_START_ROW:
                break
            continue
        if not cond_type:
            continue

        kind = cond_type.lower()
        if kind not in pending:
            continue
        intra_op, group_inv = _parse_output_cond_meta(row)
        signals = _read_signal_cells_row(row, COND_META_COLS + 1, COND_SIGNAL_MAX_COLS)
        pending[kind].append((signals, group_inv, intra_op))

    if current_output:
        flush(current_output)


def _normalize_input_mode(raw: Any) -> str:
    key = _cell_str(raw).lower()
    return INPUT_MODE_MAP.get(key, "depends")


def _parse_input_side_rows(
    rows: list[tuple[str, str, list[tuple[str, bool, str]], bool, str]],
) -> dict[str, Any]:
    """Build hi_* or lo_* fields for InputWaveSpec dict from rows of same side."""
    if not rows:
        return {}
    mode = _normalize_input_mode(rows[0][0])
    wave = rows[0][1] or "0"
    out: dict[str, Any] = {}
    if mode == "constant_0":
        out["mode"] = "constant_0"
    elif mode == "constant_1":
        out["mode"] = "constant_1"
    elif mode == "custom":
        out["mode"] = "custom"
        out["wave"] = wave
    else:
        out["mode"] = "depends"
        groups: list[list[str]] = []
        inv_g: list[list[bool]] = []
        use_g: list[list[str]] = []
        ginv: list[bool] = []
        intra: list[str] = []
        for _mode, _wave, signals, row_ginv, intra_op in rows:
            if not signals:
                continue
            groups.append([s[0] for s in signals])
            inv_g.append([s[1] for s in signals])
            use_g.append([s[2] for s in signals])
            ginv.append(row_ginv)
            intra.append(intra_op)
        if groups:
            out["groups"] = groups
            out["inv_groups"] = inv_g
            out["use_groups"] = use_g
            out["group_inv"] = ginv
            out["intra_op"] = intra
    return out


def _parse_input_conditions(rows: list[tuple], input_names: list[str]) -> dict[str, dict]:
    """Return per-input timing spec dicts."""
    current_input = ""
    hi_rows: list[tuple[str, str, list[tuple[str, bool, str]], bool, str]] = []
    lo_rows: list[tuple[str, str, list[tuple[str, bool, str]], bool, str]] = []
    specs: dict[str, dict] = {}

    def flush_input(name: str) -> None:
        if not name:
            return
        spec: dict[str, Any] = {}
        hi = _parse_input_side_rows(hi_rows)
        lo = _parse_input_side_rows(lo_rows)
        if hi:
            spec["hi_mode"] = hi["mode"]
            if hi.get("wave") is not None:
                spec["hi_wave"] = hi["wave"]
            if hi.get("groups"):
                spec["hi_groups"] = hi["groups"]
                spec["hi_inv_groups"] = hi["inv_groups"]
                spec["hi_use_groups"] = hi["use_groups"]
                spec["hi_group_inv"] = hi.get("group_inv") or []
                spec["hi_intra_op"] = hi.get("intra_op") or []
        if lo:
            spec["lo_mode"] = lo["mode"]
            if lo.get("wave") is not None:
                spec["lo_wave"] = lo["wave"]
            if lo.get("groups"):
                spec["lo_groups"] = lo["groups"]
                spec["lo_inv_groups"] = lo["inv_groups"]
                spec["lo_use_groups"] = lo["use_groups"]
                spec["lo_group_inv"] = lo.get("group_inv") or []
                spec["lo_intra_op"] = lo.get("intra_op") or []
        if spec:
            specs[name] = spec
        hi_rows.clear()
        lo_rows.clear()

    for idx, row in enumerate(rows):
        row_num = DATA_START_ROW + idx
        name_cell = _cell_str(_row_val(row, 1))
        side = _cell_str(_row_val(row, 2)).lower()
        mode_raw = _row_val(row, 3)
        wave = _cell_str(_row_val(row, 4))
        if name_cell:
            if current_input:
                flush_input(current_input)
            current_input = name_cell
        if not current_input and not side:
            if row_num > DATA_START_ROW:
                break
            continue
        if side not in ("hi", "lo"):
            continue
        signals = _read_signal_cells_row(row, INPUT_META_COLS + 1, INPUT_COND_SIGNAL_MAX_COLS)
        intra_op, row_ginv = _parse_input_cond_meta(row)
        row_tuple = (_cell_str(mode_raw), wave, signals, row_ginv, intra_op)
        if side == "hi":
            hi_rows.append(row_tuple)
        else:
            lo_rows.append(row_tuple)

    if current_input:
        flush_input(current_input)

    for name in input_names:
        if name not in specs:
            specs[name] = {"hi_mode": "depends", "lo_mode": "constant_0"}
    return specs


def _open_excel_workbook(path: str):
    """Open workbook for import (data_only; not read_only — random cell access is slow there)."""
    return load_workbook(path, data_only=True)


def load_timing_scenario_from_excel(path: str) -> TimingScenario:
    """Load TimingScenario from Excel Input Conditions + Config timing fields."""
    from timing_sim import TimingScenario

    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    wb = _open_excel_workbook(path)
    try:
        ws_config = _find_sheet(wb, "key")
        config_lookup = _config_lookup(
            _sheet_rows(ws_config, min_row=1, max_row=CONFIG_MAX_ROW, max_col=CONFIG_MAX_COL)
        )
        defaults = _parse_config_defaults(config_lookup)
        input_names: list[str] = []
        try:
            ws_nodes = _find_sheet(wb, "name")
            node_rows = _data_sheet_rows(ws_nodes, max_col=NODES_MAX_COL)
            input_names = [
                r.name for r in _parse_nodes(node_rows, defaults) if r.seq_type == "input"
            ]
        except ValueError:
            pass
        ws_in = _find_sheet(wb, "input_name")
        in_rows = _data_sheet_rows(ws_in, max_col=INPUT_COND_MAX_COL)
        inputs = _parse_input_conditions(in_rows, input_names)
        if not inputs:
            raise ValueError("No Input Conditions found in workbook")
        return TimingScenario.from_dict(
            {
                **_timing_scenario_from_config_defaults(defaults),
                "inputs": inputs,
            }
        )
    finally:
        wb.close()


def load_powerseq_from_excel(path: str) -> PowerSeqConfig:
    """Load PowerSeqConfig from Excel template workbook."""
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    wb = _open_excel_workbook(path)
    try:
        ws_config = _find_sheet(wb, "key")
        config_lookup = _config_lookup(
            _sheet_rows(ws_config, min_row=1, max_row=CONFIG_MAX_ROW, max_col=CONFIG_MAX_COL)
        )
        defaults = _parse_config_defaults(config_lookup)
        ws_nodes = _find_sheet(wb, "name")
        node_rows = _data_sheet_rows(ws_nodes, max_col=NODES_MAX_COL)
        rails = _parse_nodes(node_rows, defaults)
        rails_by_name = {r.name: r for r in rails}

        try:
            ws_out = _find_sheet(wb, "output_name")
            out_rows = _data_sheet_rows(ws_out, max_col=OUTPUT_COND_MAX_COL)
            _parse_output_conditions(out_rows, rails_by_name)
        except ValueError:
            pass

        input_names = [r.name for r in rails if r.seq_type == "input"]
        timing_scenario = _timing_scenario_from_config_defaults(defaults)
        try:
            ws_in = _find_sheet(wb, "input_name")
            in_rows = _data_sheet_rows(ws_in, max_col=INPUT_COND_MAX_COL)
            inputs = _parse_input_conditions(in_rows, input_names)
            if inputs:
                from config_models import apply_input_wave_dict

                for name, spec in inputs.items():
                    rail = rails_by_name.get(name)
                    if rail is not None and rail.seq_type == "input":
                        apply_input_wave_dict(rail, spec)
        except ValueError:
            pass

        return PowerSeqConfig(
            rails=rails,
            module_name=defaults["module_name"],
            pulses=defaults["pulses"],
            timing_scenario=timing_scenario,
        )
    finally:
        wb.close()
