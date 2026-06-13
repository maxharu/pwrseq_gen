"""
Shared Excel template header layout (Nodes sheet row 1–3).
Used by scripts/generate_excel_template.py and excel_export on save.
"""
from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def node_sheet_headers() -> list[tuple[str, str, str]]:
    return [
        ("name", "Name", "Unique;\nno duplicates"),
        ("type", "Type", "Input or Output"),
        ("cycle_hi", "CYCLE_HI", "Output: timing Hi;\nInput: DEB CYCLE_HI"),
        ("cycle_lo", "CYCLE_LO", "Output: timing Lo;\nInput: DEB CYCLE_LO"),
        ("cycle_force", "CYCLE_FORCE", "Output only"),
        ("init", "INIT", "Output: INIT;\nInput: DEB INIT"),
        ("force_val", "FORCE Value", "FORCE output 0/1;\nOutput only"),
        ("pulse_hi", "Pulse_Hi", "Pulse dropdown"),
        ("pulse_lo", "Pulse_Lo", "Pulse dropdown"),
        (
            "pulse_timing",
            "Pulse_Force / Pulse_Deb",
            "Output -> Pulse_Force;\nInput -> Pulse_Deb",
        ),
    ]


def _cell_text_width(val) -> int:
    if val is None:
        return 0
    text = str(val)
    if "\n" in text:
        return max(len(line) for line in text.split("\n"))
    return len(text)


def autosize_columns(
    ws,
    max_col: int,
    *,
    min_row: int = 1,
    max_row: int | None = None,
    min_width: int = 9,
    max_width: int = 28,
) -> None:
    row_hi = max_row if max_row is not None else ws.max_row
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in range(min_row, row_hi + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                best = max(best, min(_cell_text_width(val) + 2, max_width))
        ws.column_dimensions[letter].width = best


def apply_nodes_sheet_header_rows(ws) -> None:
    """Write Nodes sheet rows 1–3 (key / label / hint) with wrap on hint row."""
    headers = node_sheet_headers()
    ncol = len(headers)
    for c, (key, label, hint) in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=key)
        ws.cell(row=2, column=c, value=label)
        ws.cell(row=3, column=c, value=hint)
    for col in range(1, ncol + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.cell(row=2, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    autosize_columns(ws, ncol, max_row=3)
