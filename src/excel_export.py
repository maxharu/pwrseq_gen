"""
Export PowerSeqConfig to Excel node template (.xlsx / .xlsm).
"""
from __future__ import annotations

import os
import sys
import tempfile
import zipfile
from copy import copy
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.workbook.defined_name import DefinedName
except ImportError as e:
    raise ImportError("Excel export requires openpyxl: pip install openpyxl") from e

from config_models import (
    DEFAULT_PULSE,
    PowerRail,
    PowerSeqConfig,
    build_timing_scenario,
    normalize_pulse_name,
    rail_input_wave_spec,
)
from group_logic import intra_op_label
from excel_template_layout import apply_nodes_sheet_header_rows
from excel_import import (
    COND_META_COLS,
    COND_SIGNAL_MAX_COLS,
    DATA_START_ROW,
    INPUT_COND_SIGNAL_MAX_COLS,
    INPUT_META_COLS,
    MAX_DATA_ROW_CAP,
    OUTPUT_USE_SUFFIXES,
    _find_sheet,
    build_signal_dropdown_entries,
    data_area_end_row,
    last_used_row,
)
from timing_sim import DEP_HIGH, DEP_LOW, InputWaveSpec, TimingScenario
SHEET_LISTS = "Lists"

INPUT_FILL = PatternFill("solid", fgColor="F3E8FF")
OUTPUT_FILL = PatternFill("solid", fgColor="DBEAFE")
HI_FILL = PatternFill("solid", fgColor="D1FAE5")
LO_FILL = PatternFill("solid", fgColor="FEE2E2")
FORCE_FILL = PatternFill("solid", fgColor="FEF3C7")

USE_SUFFIX = {"hi": "|Hi Cond", "lo": "|Lo Cond", "force": "|Force Cond"}
INPUT_MODE_LABELS = {
    "constant_0": "Low (0)",
    "constant_1": "High (1)",
    "custom": "Custom wave",
    "depends": "Signal cond.",
}

_PKG_DIR = os.path.dirname(os.path.abspath(__file__))
_ROOT = os.path.dirname(_PKG_DIR)
DEFAULT_TEMPLATE_XLSX = os.path.join(_ROOT, "templates", "powerseq_nodes_template.xlsx")
VBA_BIN = os.path.join(_ROOT, "templates", "vba", "vbaProject.bin")

_TEMPLATE_CANDIDATES = (
    DEFAULT_TEMPLATE_XLSX,
    os.path.join(_ROOT, "templates", "powerseq_nodes_template_new.xlsm"),
    os.path.join(_ROOT, "templates", "powerseq_nodes_template_fixed.xlsm"),
    os.path.join(_ROOT, "templates", "powerseq_nodes_template.xlsm"),
)

_XLSM_GOLDEN_CANDIDATES = (
    os.path.join(_ROOT, "templates", "powerseq_nodes_template_new.xlsm"),
    os.path.join(_ROOT, "templates", "powerseq_nodes_template_fixed.xlsm"),
)


def _is_valid_workbook_template(path: str) -> bool:
    return os.path.isfile(path) and zipfile.is_zipfile(path)


def _resolve_template(template_path: str | None) -> str:
    if template_path and _is_valid_workbook_template(template_path):
        return template_path
    for candidate in _TEMPLATE_CANDIDATES:
        if _is_valid_workbook_template(candidate):
            return candidate
    raise FileNotFoundError(
        "Excel template not found (expected templates/powerseq_nodes_template.xlsx)"
    )


def _resolve_xlsm_golden() -> str | None:
    """xlsm with Sync button + drawing parts (openpyxl save strips these)."""
    for candidate in _XLSM_GOLDEN_CANDIDATES:
        if not _is_valid_workbook_template(candidate):
            continue
        try:
            with zipfile.ZipFile(candidate) as zf:
                if "xl/drawings/drawing1.xml" in zf.namelist():
                    return candidate
        except zipfile.BadZipFile:
            continue
    return None


def _finalize_xlsm(tmp_xlsx: str, xlsm_path: str) -> None:
    """Embed VBA + Sync button; re-bind sheet modules so Config pulse edits sync."""
    if not os.path.isfile(VBA_BIN):
        raise FileNotFoundError(f"VBA project not found: {VBA_BIN}")
    scripts = os.path.join(_ROOT, "scripts")
    if scripts not in sys.path:
        sys.path.insert(0, scripts)

    try:
        from embed_excel_vba import embed_vba

        embed_vba(tmp_xlsx, xlsm_path)
        return
    except Exception:
        pass

    from inject_vba_zip import copy_sync_button_from_golden, inject_vba_zip

    inject_vba_zip(tmp_xlsx, xlsm_path, VBA_BIN)
    golden = _resolve_xlsm_golden()
    if golden:
        copy_sync_button_from_golden(xlsm_path, golden)

    try:
        from embed_excel_vba import refresh_vba_inplace

        refresh_vba_inplace(xlsm_path)
        return
    except Exception:
        pass

    try:
        from embed_excel_vba import add_sync_button_inplace

        add_sync_button_inplace(xlsm_path)
    except Exception:
        pass


def format_signal_cell(name: str, inv: bool, use: str) -> str:
    if name == DEP_HIGH:
        text = "High"
    elif name == DEP_LOW:
        text = "Low"
    else:
        suffix = USE_SUFFIX.get(use, "") if use != "self" else ""
        text = f"{name}{suffix}"
    return f"!{text}" if inv else text


def _set_config_value(ws, key: str, value: Any) -> None:
    target = key.lower()
    for r in range(1, 60):
        if str(ws.cell(r, 1).value or "").strip().lower() == target:
            ws.cell(r, 4, value=value)
            return


def _clear_data_area(ws, end_col: int, *, new_row_count: int) -> None:
    clear_end = data_area_end_row(ws, new_row_count=new_row_count, max_col=end_col)
    to_unmerge = [
        str(m)
        for m in list(ws.merged_cells.ranges)
        if m.min_row >= DATA_START_ROW and m.min_col == 1
    ]
    for ref in to_unmerge:
        ws.unmerge_cells(ref)
    for r in range(DATA_START_ROW, clear_end + 1):
        for c in range(1, end_col + 1):
            cell = ws.cell(r, c)
            cell.value = None


def _write_config(ws, config: PowerSeqConfig, scenario: TimingScenario) -> None:
    _set_config_value(ws, "module_name", config.module_name)
    _set_config_value(ws, "pulses", ",".join(config.pulses))
    _set_config_value(ws, "timing_steps", scenario.steps)
    _set_config_value(ws, "timing_hscale", scenario.hscale)


def _write_nodes(ws, rails: list[PowerRail]) -> None:
    apply_nodes_sheet_header_rows(ws)
    _clear_data_area(ws, 10, new_row_count=len(rails))
    for i, rail in enumerate(rails):
        r = DATA_START_ROW + i
        fill = INPUT_FILL if rail.seq_type == "input" else OUTPUT_FILL
        if rail.seq_type == "input":
            values = (
                rail.name,
                "Input",
                rail.deb_cycle_hi,
                rail.deb_cycle_lo,
                None,
                rail.deb_init,
                None,
                rail.pulse_hi,
                rail.pulse_lo,
                rail.deb_pulse,
            )
        else:
            values = (
                rail.name,
                "Output",
                rail.cycle_hi,
                rail.cycle_lo,
                rail.cycle_force,
                rail.init,
                rail.force_val,
                rail.pulse_hi,
                rail.pulse_lo,
                rail.pulse_force,
            )
        for c, val in enumerate(values, start=1):
            cell = ws.cell(r, c, value=val)
            cell.fill = copy(fill)


def _cond_type_fill(cond_type: str | None) -> PatternFill | None:
    if not cond_type:
        return None
    key = cond_type.strip().lower()
    if key == "hi":
        return HI_FILL
    if key == "lo":
        return LO_FILL
    if key == "force":
        return FORCE_FILL
    return None


def _output_cond_rows(rail: PowerRail) -> list[tuple[str, str, str, str, list[str]]]:
    rows: list[tuple[str, str, str, str, list[str]]] = []
    block_started = False
    accessors = {
        "hi": (
            rail.get_hi_groups,
            rail.get_hi_inv,
            rail.get_hi_use,
            rail.get_hi_group_inv,
            rail.get_hi_intra_op,
            "Hi",
        ),
        "lo": (
            rail.get_lo_groups,
            rail.get_lo_inv,
            rail.get_lo_use,
            rail.get_lo_group_inv,
            rail.get_lo_intra_op,
            "Lo",
        ),
        "force": (
            rail.get_force_groups,
            rail.get_force_inv,
            rail.get_force_use,
            rail.get_force_group_inv,
            rail.get_force_intra_op,
            "Force",
        ),
    }
    for kind in ("hi", "lo", "force"):
        groups_fn, get_inv, get_use, get_ginv, get_intra, label = accessors[kind]
        groups = groups_fn()
        non_empty = [g for g in groups if g]
        if not non_empty:
            rows.append((rail.name if not block_started else "", label, "AND", "N", []))
            block_started = True
            continue
        for gi, group in enumerate(groups):
            if not group:
                continue
            signals = [
                format_signal_cell(n, get_inv(gi, ii, n), get_use(gi, ii, n))
                for ii, n in enumerate(group)
            ]
            ginv = "Y" if get_ginv(gi) else "N"
            op = intra_op_label(get_intra(gi))
            rows.append((rail.name if not block_started else "", label, op, ginv, signals))
            block_started = True
    return rows


def _write_output_conditions(ws, config: PowerSeqConfig) -> None:
    sig_end = COND_META_COLS + COND_SIGNAL_MAX_COLS
    all_rows: list[tuple[str, str, str, str, list[str]]] = []
    for rail in config.rails:
        if rail.seq_type != "output":
            continue
        all_rows.extend(_output_cond_rows(rail))
    _clear_data_area(ws, sig_end, new_row_count=len(all_rows))

    block_starts: list[tuple[int, int]] = []
    block_start = DATA_START_ROW
    for i, (output_name, cond_type, intra_op, group_inv, signals) in enumerate(all_rows):
        r = DATA_START_ROW + i
        if output_name:
            if i > 0:
                block_starts.append((block_start, r - 1))
            block_start = r
        ws.cell(r, 1, value=output_name or None)
        ws.cell(r, 2, value=cond_type)
        ws.cell(r, 3, value=intra_op)
        ws.cell(r, 4, value=group_inv)
        fill = _cond_type_fill(cond_type)
        for si, sig in enumerate(signals, start=1):
            ws.cell(r, COND_META_COLS + si, value=sig)
        if fill:
            for c in range(1, sig_end + 1):
                ws.cell(r, c).fill = copy(fill)
    if all_rows:
        block_starts.append((block_start, DATA_START_ROW + len(all_rows) - 1))
    for start, end in block_starts:
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)


def _side_rows(spec: InputWaveSpec, side: str) -> list[tuple[str, str | None, str, str, list[str]]]:
    mode = getattr(spec, f"{side}_mode")
    if mode == "constant_0":
        return [(INPUT_MODE_LABELS["constant_0"], None, "AND", "N", [])]
    if mode == "constant_1":
        return [(INPUT_MODE_LABELS["constant_1"], None, "AND", "N", [])]
    if mode == "custom":
        wave = getattr(spec, f"{side}_wave", "0") or "0"
        return [(INPUT_MODE_LABELS["custom"], wave, "AND", "N", [])]
    groups = getattr(spec, f"{side}_groups") or []
    inv_groups = getattr(spec, f"{side}_inv_groups") or []
    use_groups = getattr(spec, f"{side}_use_groups") or []
    group_inv_list = getattr(spec, f"{side}_group_inv") or []
    intra_op_list = getattr(spec, f"{side}_intra_op") or []
    non_empty = [g for g in groups if g]
    if not non_empty:
        return [(INPUT_MODE_LABELS["depends"], None, "AND", "N", [])]
    rows: list[tuple[str, str | None, str, str, list[str]]] = []
    for gi, group in enumerate(groups):
        if not group:
            continue
        inv_g = inv_groups[gi] if gi < len(inv_groups) else []
        use_g = use_groups[gi] if gi < len(use_groups) else []
        ginv = "Y" if (gi < len(group_inv_list) and group_inv_list[gi]) else "N"
        op = intra_op_label(intra_op_list[gi] if gi < len(intra_op_list) else "and")
        signals = []
        for ii, n in enumerate(group):
            inv = inv_g[ii] if ii < len(inv_g) else False
            use = use_g[ii] if ii < len(use_g) else "self"
            signals.append(format_signal_cell(n, inv, use))
        rows.append((INPUT_MODE_LABELS["depends"], None, op, ginv, signals))
    return rows


def _input_cond_rows(name: str, spec: InputWaveSpec) -> list[tuple[str, str, str, str | None, str, str, list[str]]]:
    rows: list[tuple[str, str, str, str | None, str, str, list[str]]] = []
    block_started = False
    for side_label, side in (("Hi", "hi"), ("Lo", "lo")):
        for mode, wave, op, ginv, signals in _side_rows(spec, side):
            rows.append(
                (
                    name if not block_started else "",
                    side_label,
                    mode,
                    wave,
                    op,
                    ginv,
                    signals,
                )
            )
            block_started = True
    if not rows:
        rows = [
            (name, "Hi", INPUT_MODE_LABELS["depends"], None, "AND", "N", []),
            ("", "Lo", INPUT_MODE_LABELS["constant_0"], None, "AND", "N", []),
        ]
    return rows


def _write_input_conditions(ws, config: PowerSeqConfig, scenario: TimingScenario) -> None:
    sig_end = INPUT_META_COLS + INPUT_COND_SIGNAL_MAX_COLS
    all_rows: list[tuple[str, str, str, str | None, str, str, list[str]]] = []
    for rail in config.rails:
        if rail.seq_type != "input":
            continue
        spec = rail_input_wave_spec(rail)
        all_rows.extend(_input_cond_rows(rail.name, spec))
    _clear_data_area(ws, sig_end, new_row_count=len(all_rows))

    block_starts: list[tuple[int, int]] = []
    block_start = DATA_START_ROW
    for i, (input_name, side, mode, wave, operation, group_inv, signals) in enumerate(all_rows):
        r = DATA_START_ROW + i
        if input_name:
            if i > 0:
                block_starts.append((block_start, r - 1))
            block_start = r
        ws.cell(r, 1, value=input_name or None)
        ws.cell(r, 2, value=side)
        ws.cell(r, 3, value=mode)
        ws.cell(r, 4, value=wave)
        ws.cell(r, 5, value=operation)
        ws.cell(r, 6, value=group_inv)
        fill = HI_FILL if side == "Hi" else LO_FILL
        for si, sig in enumerate(signals, start=1):
            ws.cell(r, INPUT_META_COLS + si, value=sig)
        for c in range(1, sig_end + 1):
            ws.cell(r, c).fill = copy(fill)
    if all_rows:
        block_starts.append((block_start, DATA_START_ROW + len(all_rows) - 1))
    for start, end in block_starts:
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)


def _collect_signal_list(config: PowerSeqConfig, cond_signals: set[str]) -> list[str]:
    nodes = [(rail.name, rail.seq_type) for rail in config.rails]
    return build_signal_dropdown_entries(nodes, extra=sorted(cond_signals))


def _gather_cond_signal_names(config: PowerSeqConfig) -> set[str]:
    names: set[str] = set()

    def add_from_rail(rail: PowerRail) -> None:
        for kind in ("hi", "lo", "force"):
            groups_fn = getattr(rail, f"get_{kind}_groups")
            get_use = getattr(rail, f"get_{kind}_use")
            for gi, group in enumerate(groups_fn()):
                for ii, n in enumerate(group):
                    if n not in (DEP_HIGH, DEP_LOW):
                        names.add(n)
                    use = get_use(gi, ii, n)
                    if use != "self" and n not in (DEP_HIGH, DEP_LOW):
                        names.add(n)

    for rail in config.rails:
        if rail.seq_type == "output":
            add_from_rail(rail)

    return names


def _write_lists(wb, config: PowerSeqConfig, cond_signals: set[str]) -> None:
    if SHEET_LISTS not in wb.sheetnames:
        return
    ws = wb[SHEET_LISTS]
    entries = _collect_signal_list(config, cond_signals)
    pulses = config.pulses or [DEFAULT_PULSE]
    write_end = max(len(entries) + 1, len(pulses) + 1)
    used_end = last_used_row(ws, min_row=2, max_col=2)
    max_row = min(max(used_end, write_end), MAX_DATA_ROW_CAP)
    for r in range(2, max_row + 1):
        ws.cell(r, 1, value=None)
        ws.cell(r, 2, value=None)
    for i, name in enumerate(entries, start=2):
        ws.cell(i, 1, value=name)
    for i, name in enumerate(pulses, start=2):
        ws.cell(i, 2, value=normalize_pulse_name(name))
    sig_last = len(entries) + 1
    pulse_last = max(len(pulses) + 1, 2)
    for name in ("SignalList", "PulseList"):
        if name in wb.defined_names:
            del wb.defined_names[name]
    wb.defined_names.add(
        DefinedName("SignalList", attr_text=f"{SHEET_LISTS}!$A$2:$A${sig_last}")
    )
    wb.defined_names.add(
        DefinedName("PulseList", attr_text=f"{SHEET_LISTS}!$B$2:$B${pulse_last}")
    )


def _resolve_scenario(config: PowerSeqConfig) -> TimingScenario:
    embedded = config.timing_scenario or {}
    if embedded.get("inputs"):
        from timing_scenario_io import merge_scenario_for_config

        return merge_scenario_for_config(TimingScenario.from_dict(embedded), config)
    return build_timing_scenario(config)


def export_powerseq_to_excel(
    config: PowerSeqConfig,
    path: str,
    *,
    template_path: str | None = None,
) -> None:
    """Write config to Excel template format (.xlsx or .xlsm)."""
    out_ext = os.path.splitext(path)[1].lower()
    if template_path:
        tpl = _resolve_template(template_path)
    elif out_ext == ".xlsm":
        tpl = _resolve_xlsm_golden() or _resolve_template(None)
    else:
        tpl = _resolve_template(None)
    wb = load_workbook(tpl)
    scenario = _resolve_scenario(config)
    cond_signals = _gather_cond_signal_names(config)

    _write_config(_find_sheet(wb, "key"), config, scenario)
    _write_nodes(_find_sheet(wb, "name"), config.rails)
    try:
        _write_output_conditions(_find_sheet(wb, "output_name"), config)
    except ValueError:
        pass
    try:
        _write_input_conditions(_find_sheet(wb, "input_name"), config, scenario)
    except ValueError:
        pass
    _write_lists(wb, config, cond_signals)

    if out_ext == ".xlsm":
        tmp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
        try:
            wb.save(tmp_xlsx)
            _finalize_xlsm(tmp_xlsx, path)
        finally:
            wb.close()
            if os.path.isfile(tmp_xlsx):
                os.unlink(tmp_xlsx)
    else:
        wb.save(path)
        wb.close()
